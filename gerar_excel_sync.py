#!/usr/bin/env python3
"""
Script para gerar relatório Excel com sincronização automática
"""
import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def carregar_dados():
    """Carrega dados dos arquivos JSON"""
    dados = {
        'movimentacoes': [],
        'ferramentas': [],
        'solicitantes': []
    }

    arquivos = {
        'movimentacoes': 'movimentacoes.json',
        'ferramentas': 'ferramentas.json',
        'solicitantes': 'solicitantes.json'
    }

    for chave, arquivo in arquivos.items():
        caminho = os.path.join(os.path.dirname(__file__), arquivo)
        if os.path.exists(caminho):
            try:
                with open(caminho, 'r', encoding='utf-8') as f:
                    dados[chave] = json.load(f)
            except Exception as e:
                print(f"Erro ao carregar {arquivo}: {e}")
                dados[chave] = []

    return dados

def criar_planilha_retiradas(wb, movimentacoes):
    """Cria a planilha de retiradas"""
    ws = wb.create_sheet("Retiradas")

    # Cabeçalhos
    headers = ['Data', 'Ferramenta', 'Patrimônio', 'Solicitante', 'Tipo', 'Data Devolução', 'Hora Devolução', 'Tem Retorno', 'Observações']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Dados
    for row, mov in enumerate(movimentacoes, 2):
        ws.cell(row=row, column=1, value=mov.get('dataRetirada', ''))
        ws.cell(row=row, column=2, value=mov.get('ferramenta', ''))
        ws.cell(row=row, column=3, value=mov.get('patrimonio', ''))
        ws.cell(row=row, column=4, value=mov.get('solicitante', ''))
        ws.cell(row=row, column=5, value=mov.get('tipo', ''))
        ws.cell(row=row, column=6, value=mov.get('dataRetorno', ''))
        ws.cell(row=row, column=7, value=mov.get('horaRetorno', ''))
        ws.cell(row=row, column=8, value=mov.get('temRetorno', ''))
        ws.cell(row=row, column=9, value=mov.get('observacao', ''))

    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

def criar_planilha_estoque(wb, ferramentas, solicitantes):
    """Cria a planilha de estoque/administrativo"""
    ws = wb.create_sheet("Estoque")

    # Cabeçalhos
    headers = ['Tipo', 'Nome', 'Patrimônios', 'Descrição']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    row = 2
    # Ferramentas
    for ferramenta in ferramentas:
        ws.cell(row=row, column=1, value='Ferramenta')
        ws.cell(row=row, column=2, value=ferramenta.get('nome', ''))
        patrimonios = ferramenta.get('patrimonios', [])
        ws.cell(row=row, column=3, value=', '.join(patrimonios) if patrimonios else '')
        ws.cell(row=row, column=4, value=f'Quantidade: {len(patrimonios)}')
        row += 1

    # Solicitantes
    for solicitante in solicitantes:
        ws.cell(row=row, column=1, value='Solicitante')
        ws.cell(row=row, column=2, value=solicitante.get('nome', ''))
        ws.cell(row=row, column=3, value='')
        ws.cell(row=row, column=4, value='')
        row += 1

    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

def adicionar_vba_sincronizacao(wb):
    """Adiciona código VBA para sincronização automática"""
    vba_code = '''
Private Sub Workbook_Open()
    On Error Resume Next
    Call SincronizarDados
End Sub

Sub SincronizarDados()
    On Error GoTo ErrorHandler

    ' Verificar se já sincronizou hoje
    If Sheets("Config").Range("B1").Value = Format(Date, "yyyy-mm-dd") Then
        Exit Sub
    End If

    ' Mostrar progresso
    Application.StatusBar = "Sincronizando dados..."
    Application.ScreenUpdating = False

    ' Obter timestamp da última sincronização
    Dim lastSync As String
    lastSync = Sheets("Config").Range("B2").Value

    ' URL do servidor local
    Dim serverUrl As String
    serverUrl = "http://localhost:8000/api/sync"

    If lastSync <> "" Then
        serverUrl = serverUrl & "?last_sync=" & lastSync
    End If

    ' Fazer request HTTP
    Dim http As Object
    Set http = CreateObject("MSXML2.XMLHTTP")

    http.Open "GET", serverUrl, False
    http.setRequestHeader "Content-Type", "application/json"
    http.send

    If http.Status = 200 Then
        ' Parse JSON response
        Dim jsonText As String
        jsonText = http.responseText

        ' Usar ScriptControl para parse JSON (ou alternativa)
        Call ProcessarDadosJSON(jsonText)

        ' Atualizar timestamp
        Sheets("Config").Range("B2").Value = Format(Now, "yyyy-mm-ddTHH:mm:ssZ")
        Sheets("Config").Range("B1").Value = Format(Date, "yyyy-mm-dd")

        MsgBox "Dados sincronizados com sucesso!", vbInformation
    Else
        MsgBox "Erro ao sincronizar dados. Verifique se o servidor está rodando.", vbExclamation
    End If

    Application.StatusBar = False
    Application.ScreenUpdating = True
    Exit Sub

ErrorHandler:
    Application.StatusBar = False
    Application.ScreenUpdating = True
    MsgBox "Erro durante sincronização: " & Err.Description, vbCritical
End Sub

Sub ProcessarDadosJSON(jsonText As String)
    ' Esta é uma implementação simplificada
    ' Em produção, você precisaria de uma biblioteca JSON completa

    ' Para este exemplo, vamos assumir que o JSON é simples
    ' e processar manualmente

    Dim movimentacoes As String
    Dim ferramentas As String
    Dim solicitantes As String
    Dim timestamp As String

    ' Extrair dados do JSON (simplificado)
    movimentacoes = ExtrairValorJSON(jsonText, "movimentacoes")
    ferramentas = ExtrairValorJSON(jsonText, "ferramentas")
    solicitantes = ExtrairValorJSON(jsonText, "solicitantes")
    timestamp = ExtrairValorJSON(jsonText, "timestamp")

    ' Processar movimentações
    If movimentacoes <> "" Then
        Call AdicionarMovimentacoes(movimentacoes)
    End If

    ' Processar ferramentas
    If ferramentas <> "" Then
        Call AdicionarFerramentas(ferramentas)
    End If

    ' Processar solicitantes
    If solicitantes <> "" Then
        Call AdicionarSolicitantes(solicitantes)
    End If
End Sub

Function ExtrairValorJSON(json As String, chave As String) As String
    ' Função simplificada para extrair valores de JSON
    Dim pos As Long
    pos = InStr(json, """" & chave & """:")

    If pos > 0 Then
        Dim startPos As Long
        startPos = pos + Len(chave) + 3

        ' Encontrar o fim do valor (simplificado)
        Dim endPos As Long
        Dim bracketCount As Long
        Dim inString As Boolean
        Dim i As Long

        For i = startPos To Len(json)
            Select Case Mid(json, i, 1)
                Case """"
                    inString = Not inString
                Case "[", "{"
                    If Not inString Then bracketCount = bracketCount + 1
                Case "]", "}"
                    If Not inString Then bracketCount = bracketCount - 1
                Case ","
                    If Not inString And bracketCount = 0 Then
                        endPos = i - 1
                        Exit For
                    End If
            End Select
        Next i

        If endPos = 0 Then endPos = Len(json)

        ExtrairValorJSON = Mid(json, startPos, endPos - startPos + 1)
    Else
        ExtrairValorJSON = ""
    End If
End Function

Sub AdicionarMovimentacoes(jsonArray As String)
    Dim ws As Worksheet
    Set ws = Sheets("Retiradas")

    ' Encontrar última linha
    Dim lastRow As Long
    lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row + 1

    ' Parse array JSON simplificado
    Dim items() As String
    items = Split(Replace(Replace(jsonArray, "[", ""), "]", ""), "},{")

    Dim i As Long
    For i = LBound(items) To UBound(items)
        Dim item As String
        item = Replace(Replace(items(i), "{", ""), "}", "")

        If i = LBound(items) And Left(item, 1) <> """" Then item = "{" & item
        If i = UBound(items) And Right(item, 1) <> """" Then item = item & "}"

        ' Extrair valores
        Dim dataRetirada As String
        Dim ferramenta As String
        Dim patrimonio As String
        Dim solicitante As String
        Dim tipo As String
        Dim dataRetorno As String
        Dim horaRetorno As String
        Dim temRetorno As String
        Dim observacao As String

        dataRetirada = ExtrairValorJSON(item, "dataRetirada")
        ferramenta = ExtrairValorJSON(item, "ferramenta")
        patrimonio = ExtrairValorJSON(item, "patrimonio")
        solicitante = ExtrairValorJSON(item, "solicitante")
        tipo = ExtrairValorJSON(item, "tipo")
        dataRetorno = ExtrairValorJSON(item, "dataRetorno")
        horaRetorno = ExtrairValorJSON(item, "horaRetorno")
        temRetorno = ExtrairValorJSON(item, "temRetorno")
        observacao = ExtrairValorJSON(item, "observacao")

        ' Remover aspas
        dataRetirada = Replace(Replace(dataRetirada, """", ""), ",", "")
        ferramenta = Replace(Replace(ferramenta, """", ""), ",", "")
        patrimonio = Replace(Replace(patrimonio, """", ""), ",", "")
        solicitante = Replace(Replace(solicitante, """", ""), ",", "")
        tipo = Replace(Replace(tipo, """", ""), ",", "")
        dataRetorno = Replace(Replace(dataRetorno, """", ""), ",", "")
        horaRetorno = Replace(Replace(horaRetorno, """", ""), ",", "")
        temRetorno = Replace(Replace(temRetorno, """", ""), ",", "")
        observacao = Replace(Replace(observacao, """", ""), ",", "")

        ' Adicionar à planilha
        ws.Cells(lastRow, 1).Value = dataRetirada
        ws.Cells(lastRow, 2).Value = ferramenta
        ws.Cells(lastRow, 3).Value = patrimonio
        ws.Cells(lastRow, 4).Value = solicitante
        ws.Cells(lastRow, 5).Value = tipo
        ws.Cells(lastRow, 6).Value = dataRetorno
        ws.Cells(lastRow, 7).Value = horaRetorno
        ws.Cells(lastRow, 8).Value = temRetorno
        ws.Cells(lastRow, 9).Value = observacao

        lastRow = lastRow + 1
    Next i
End Sub

Sub AdicionarFerramentas(jsonArray As String)
    Dim ws As Worksheet
    Set ws = Sheets("Estoque")

    ' Encontrar última linha de ferramentas
    Dim lastRow As Long
    lastRow = 2 ' Começar da linha 2

    Do While ws.Cells(lastRow, 1).Value <> ""
        lastRow = lastRow + 1
    Loop

    ' Parse array JSON simplificado
    Dim items() As String
    items = Split(Replace(Replace(jsonArray, "[", ""), "]", ""), "},{")

    Dim i As Long
    For i = LBound(items) To UBound(items)
        Dim item As String
        item = Replace(Replace(items(i), "{", ""), "}", "")

        If i = LBound(items) And Left(item, 1) <> """" Then item = "{" & item
        If i = UBound(items) And Right(item, 1) <> """" Then item = item & "}"

        ' Extrair valores
        Dim nome As String
        Dim patrimonios As String

        nome = ExtrairValorJSON(item, "nome")
        patrimonios = ExtrairValorJSON(item, "patrimonios")

        ' Remover aspas
        nome = Replace(Replace(nome, """", ""), ",", "")
        patrimonios = Replace(Replace(patrimonios, "[", ""), "]", "")
        patrimonios = Replace(patrimonios, """", "")

        ' Adicionar à planilha
        ws.Cells(lastRow, 1).Value = "Ferramenta"
        ws.Cells(lastRow, 2).Value = nome
        ws.Cells(lastRow, 3).Value = patrimonios
        ws.Cells(lastRow, 4).Value = "Atualizado"

        lastRow = lastRow + 1
    Next i
End Sub

Sub AdicionarSolicitantes(jsonArray As String)
    Dim ws As Worksheet
    Set ws = Sheets("Estoque")

    ' Encontrar última linha
    Dim lastRow As Long
    lastRow = 2

    Do While ws.Cells(lastRow, 1).Value <> ""
        lastRow = lastRow + 1
    Loop

    ' Parse array JSON simplificado
    Dim items() As String
    items = Split(Replace(Replace(jsonArray, "[", ""), "]", ""), "},{")

    Dim i As Long
    For i = LBound(items) To UBound(items)
        Dim item As String
        item = Replace(Replace(items(i), "{", ""), "}", "")

        If i = LBound(items) And Left(item, 1) <> """" Then item = "{" & item
        If i = UBound(items) And Right(item, 1) <> """" Then item = item & "}"

        ' Extrair valores
        Dim nome As String
        nome = ExtrairValorJSON(item, "nome")

        ' Remover aspas
        nome = Replace(Replace(nome, """", ""), ",", "")

        ' Adicionar à planilha
        ws.Cells(lastRow, 1).Value = "Solicitante"
        ws.Cells(lastRow, 2).Value = nome
        ws.Cells(lastRow, 3).Value = ""
        ws.Cells(lastRow, 4).Value = "Atualizado"

        lastRow = lastRow + 1
    Next i
End Sub
'''

    # Criar planilha de configuração
    config_ws = wb.create_sheet("Config")
    config_ws.cell(row=1, column=1, value="Última Sincronização (Data)")
    config_ws.cell(row=1, column=2, value="")
    config_ws.cell(row=2, column=1, value="Última Sincronização (Timestamp)")
    config_ws.cell(row=2, column=2, value=datetime.now().isoformat())
    config_ws.cell(row=3, column=1, value="URL Servidor")
    config_ws.cell(row=3, column=2, value="http://localhost:8000")

    # Ocultar planilha de configuração
    config_ws.sheet_state = 'hidden'

    # Nota: VBA não pode ser adicionado diretamente via openpyxl
    # O código VBA precisa ser adicionado manualmente no Excel
    # ou usando uma biblioteca especializada como xlwings

    print("ATENÇÃO: Código VBA precisa ser adicionado manualmente ao Excel")
    print("Para adicionar o VBA:")
    print("1. Abra o arquivo Excel")
    print("2. Pressione Alt+F11 para abrir o editor VBA")
    print("3. Insira um módulo e cole o código VBA fornecido")
    print("4. Salve como .xlsm (Excel com macros)")

    # Salvar instruções VBA em arquivo separado
    with open('vba_code.txt', 'w', encoding='utf-8') as f:
        f.write(vba_code)

    print("Código VBA salvo em: vba_code.txt")

def gerar_relatorio():
    """Gera o relatório Excel completo"""
    print("Carregando dados...")
    dados = carregar_dados()

    print("Criando workbook...")
    wb = openpyxl.Workbook()

    # Remover planilha padrão
    wb.remove(wb.active)

    print("Criando planilhas...")
    criar_planilha_retiradas(wb, dados['movimentacoes'])
    criar_planilha_estoque(wb, dados['ferramentas'], dados['solicitantes'])

    print("Adicionando sincronização VBA...")
    adicionar_vba_sincronizacao(wb)

    # Salvar arquivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f"relatorio_sincronizado_{timestamp}.xlsm"

    print(f"Salvando arquivo: {nome_arquivo}")
    wb.save(nome_arquivo)

    print("Relatório gerado com sucesso!")
    print(f"Arquivo: {nome_arquivo}")
    print("\nPara sincronização automática:")
    print("1. Abra o arquivo Excel")
    print("2. Habilite macros se solicitado")
    print("3. Certifique-se que o servidor está rodando (python server_sync.py)")
    print("4. O Excel irá sincronizar automaticamente ao abrir")

    return nome_arquivo

if __name__ == "__main__":
    gerar_relatorio()
